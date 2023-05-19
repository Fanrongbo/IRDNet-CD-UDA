
import numpy as np
import os
import argparse as ag
import json
import shutil
import pickle
from IPython import display
import torch.nn as nn
import matplotlib.pyplot as plt
import torch.nn.init as init
from option.config import cfg
import openpyxl
import time
from shutil import copyfile
from openpyxl.styles import PatternFill
import openpyxl as op


def initialize_weights(*models):
    for model in models:
        for m in model.modules():
            if isinstance(m, nn.Conv2d):
                nn.init.kaiming_normal_(m.weight.data, nonlinearity='relu')
            elif isinstance(m, nn.BatchNorm2d):
                m.weight.data.fill_(1.)
                m.bias.data.fill_(1e-4)
            elif isinstance(m, nn.Linear):
                m.weight.data.normal_(0.0, 0.0001)
                m.bias.data.zero_()
def save_pickle(data, file_name):
    f = open(file_name, "wb")
    pickle.dump(data, f)
    f.close()


def load_pickle(file_name):
    f = open(file_name, "rb+")
    data = pickle.load(f)
    f.close()
    return data
class setFigure():
    # def __init__(self):
    def initialize_figure(self):
        self.metrics = {
            'nochange_acc': [],
            'change_acc': [],
            'prec': [],
            'rec': [],
            'f_meas': [],
            'Loss': [],
            'ce_loss': [],
            'total_acc': [],
            'DALoss': [],
            'FeatLoss':[],
            'Iou':[]

        }
        return self.metrics
    def set_figure(self,metric_dict, nochange_acc, change_acc, prec, rec, f_meas, Loss, ce_loss, total_acc,DALoss,FeatLoss,Iou):
        metric_dict['nochange_acc'].append(nochange_acc)
        metric_dict['change_acc'].append(change_acc)
        metric_dict['prec'].append(prec)
        metric_dict['rec'].append(rec)
        metric_dict['f_meas'].append(f_meas)
        metric_dict['Loss'].append(Loss)
        metric_dict['ce_loss'].append(ce_loss)
        metric_dict['total_acc'].append(total_acc)
        # metric_dict['focal_loss'].append(focal_loss)
        metric_dict['DALoss'].append(DALoss)
        # metric_dict['dice_loss'].append(dice_loss)
        metric_dict['FeatLoss'].append(FeatLoss)
        metric_dict['Iou'].append(Iou)


        return metric_dict

class setFigureDA():
    # def __init__(self):
    def initialize_figure(self):
        self.metrics = {
            'nochange_acc': [],
            'change_acc': [],
            'prec': [],
            'rec': [],
            'f_meas': [],
            'Loss': [],
            'ce_loss': [],
            'total_acc': [],
            'DAlowLoss': [],
            'FeatLoss':[],
            'Iou':[],
            'DAEntropy':[],
            'DATLoss':[]

        }
        return self.metrics
    def set_figure(self,metric_dict, nochange_acc, change_acc, prec, rec, f_meas, Loss, ce_loss, total_acc,DAlowLoss,DAEntropy,DATLoss,FeatLoss,Iou):
        metric_dict['nochange_acc'].append(nochange_acc)
        metric_dict['change_acc'].append(change_acc)
        metric_dict['prec'].append(prec)
        metric_dict['rec'].append(rec)
        metric_dict['f_meas'].append(f_meas)
        metric_dict['Loss'].append(Loss)
        metric_dict['ce_loss'].append(ce_loss)
        metric_dict['total_acc'].append(total_acc)
        metric_dict['DAEntropy'].append(DAEntropy)
        metric_dict['DAlowLoss'].append(DAlowLoss)
        metric_dict['DATLoss'].append(DATLoss)
        metric_dict['FeatLoss'].append(FeatLoss)
        metric_dict['Iou'].append(Iou)


        return metric_dict
def add_weight_decay(net, l2_value, skip_list=()):
    # https://raberrytv.wordpress.com/2017/10/29/pytorch-weight-decay-made-easy/

    decay, no_decay = [], []
    for name, param in net.named_parameters():
        if not param.requires_grad:
            continue # frozen weights
        if len(param.shape) == 1 or name.endswith(".bias") or name in skip_list:
            no_decay.append(param)
        else:
            decay.append(param)

    return [{'params': no_decay, 'weight_decay': 0.0}, {'params': decay, 'weight_decay': l2_value}]

def get_parser_with_args(metadata_json='./utils/metadata_GZ.json'):
    parser = ag.ArgumentParser(description='Training change detection network')

    with open(metadata_json, 'r') as fin:
        metadata = json.load(fin)
        parser.set_defaults(**metadata)
        return parser, metadata
    return None
def plotFigure(figure_train_metrics,figure_test_metrics,num_epochs,name,model_type, time_now):
    t = np.linspace(1, num_epochs, num_epochs)
    e=num_epochs
    # print(e,len(figure_train_metrics['rec']),len(figure_test_metrics['ce_loss']))
    epoch_train_nochange_accuracy = figure_train_metrics['nochange_acc']
    epoch_train_change_accuracy = figure_train_metrics['change_acc']
    epoch_train_precision = figure_train_metrics['prec']
    epoch_train_recall = figure_train_metrics['rec']
    epoch_train_Fmeasure = figure_train_metrics['f_meas']
    epoch_train_Loss = figure_train_metrics['Loss']
    epoch_train_ce_loss = figure_train_metrics['ce_loss']
    epoch_train_accuracy = figure_train_metrics['total_acc']
    epoch_train_DAEntropy = figure_train_metrics['DAEntropy']
    epoch_train_DAlowLoss = figure_train_metrics['DAlowLoss']
    epoch_train_FeatLoss = figure_train_metrics['FeatLoss']
    epoch_train_DATLoss = figure_train_metrics['DATLoss']
    epoch_train_Iou = figure_train_metrics['Iou']

    epoch_test_nochange_accuracy = figure_test_metrics['nochange_acc']
    epoch_test_change_accuracy = figure_test_metrics['change_acc']
    epoch_test_precision = figure_test_metrics['prec']
    epoch_test_recall = figure_test_metrics['rec']
    epoch_test_Fmeasure = figure_test_metrics['f_meas']
    epoch_test_Loss= figure_test_metrics['Loss']
    epoch_test_ce_loss= figure_test_metrics['ce_loss']
    epoch_test_accuracy = figure_test_metrics['total_acc']

    epoch_test_DAEntropy= figure_test_metrics['DAEntropy']
    epoch_test_DAlowLoss= figure_test_metrics['DAlowLoss']
    epoch_test_FeatLoss = figure_test_metrics['FeatLoss']
    epoch_test_DATLoss= figure_test_metrics['DATLoss']
    epoch_test_Iou= figure_test_metrics['Iou']

    plt.figure(num=1)
    plt.clf()

    l1_1, = plt.plot(t[:e + 1], epoch_train_ce_loss[:e + 1], label='Train CE loss')
    l1_2, = plt.plot(t[:e + 1], epoch_test_ce_loss[:e + 1], label='Test CE loss')
    l1_5, = plt.plot(t[:e + 1], epoch_train_FeatLoss[:e + 1], label='Train Feat Loss')
    l1_6, = plt.plot(t[:e + 1], epoch_test_FeatLoss[:e + 1], label='Test Feat Loss')
    l1_7, = plt.plot(t[:e + 1], epoch_train_Loss, label='Train Total loss')
    l1_8, = plt.plot(t[:e + 1], epoch_test_Loss, label='Test Total loss')
    plt.legend(handles=[l1_1, l1_2, l1_5, l1_6,l1_7,l1_8])
    plt.grid()
    #         plt.gcf().gca().set_ylim(bottom = 0)
    plt.gcf().gca().set_xlim(left=0)
    plt.title('Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=2)
    plt.clf()
    l2_1, = plt.plot(t[:e + 1], epoch_train_accuracy[:e + 1], label='Train accuracy')
    l2_2, = plt.plot(t[:e + 1], epoch_test_accuracy[:e + 1], label='Test accuracy')
    plt.legend(handles=[l2_1, l2_2])
    plt.grid()
    plt.gcf().gca().set_ylim(0, 1)
    plt.title('Accuracy')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=3)
    plt.clf()
    l3_1, = plt.plot(t[:e + 1], epoch_train_nochange_accuracy[:e + 1], label='Train accuracy: no change')
    l3_2, = plt.plot(t[:e + 1], epoch_train_change_accuracy[:e + 1], label='Train accuracy: change')
    l3_3, = plt.plot(t[:e + 1], epoch_test_nochange_accuracy[:e + 1], label='Test accuracy: no change')
    l3_4, = plt.plot(t[:e + 1], epoch_test_change_accuracy[:e + 1], label='Test accuracy: change')
    plt.legend(loc='best', handles=[l3_1, l3_2, l3_3, l3_4])
    plt.grid()
    plt.gcf().gca().set_ylim(0, 1)
    plt.title('Accuracy per class')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=4)
    plt.clf()
    l4_1, = plt.plot(t[:e + 1], epoch_train_precision[:e + 1], label='Train precision')
    l4_2, = plt.plot(t[:e + 1], epoch_train_recall[:e + 1], label='Train recall')
    l4_3, = plt.plot(t[:e + 1], epoch_train_Fmeasure[:e + 1], label='Train Dice/F1')
    l4_4, = plt.plot(t[:e + 1], epoch_test_precision[:e + 1], label='Test precision')
    l4_5, = plt.plot(t[:e + 1], epoch_test_recall[:e + 1], label='Test recall')
    l4_6, = plt.plot(t[:e + 1], epoch_test_Fmeasure[:e + 1], label='Test Dice/F1')
    l4_7, = plt.plot(t[:e + 1], epoch_train_Iou[:e + 1], label='Train Iou')
    l4_8, = plt.plot(t[:e + 1], epoch_test_Iou[:e + 1], label='Test Iou')
    plt.legend(loc='best', handles=[l4_1, l4_2, l4_3, l4_4, l4_5, l4_6,l4_7,l4_8])
    plt.grid()
    plt.gcf().gca().set_ylim(0, 1)
    #         plt.gcf().gca().set_ylim(bottom = 0)
    #         plt.gcf().gca().set_xlim(left = 0)
    plt.title('Precision, Recall , F-measure and Iou')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=5)
    plt.clf()
    l5_1, = plt.plot(t[:e + 1], epoch_train_ce_loss[:e + 1], label='Train CELoss')
    l5_2, = plt.plot(t[:e + 1], epoch_test_ce_loss[:e + 1], label='Test CELoss')
    plt.legend(loc='best', handles=[l5_1, l5_2])
    plt.grid()
    plt.gcf().gca().set_xlim(left=0)
    plt.title('CE Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=6)
    plt.clf()
    l6_1, = plt.plot(t[:e + 1], epoch_train_Loss[:e + 1], label='Train Loss')
    l6_2, = plt.plot(t[:e + 1], epoch_test_Loss[:e + 1], label='Test Loss')

    plt.legend(loc='best', handles=[l6_1,l6_2])
    plt.grid()
    plt.gcf().gca().set_xlim(left=0)
    plt.title('Total Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=7)
    plt.clf()
    train_loss = []
    test_loss = []
    # for i in range(len(epoch_train_Loss[:e + 1])):
    #     train_loss.append(epoch_train_clf_loss[:e + 1][i]+epoch_train_marginLoss[:e + 1][i])
    #     test_loss.append(epoch_test_loss[:e + 1][i]+epoch_test_marginLoss[:e + 1][i])

    l7_1, = plt.plot(t[:e + 1], epoch_train_DAEntropy[:e + 1], label='Train DAEntropy loss')
    # l7_2, = plt.plot(t[:e + 1], epoch_test_DAEntropy[:e + 1], label='Test DAEntropy loss')
    l7_3, = plt.plot(t[:e + 1], epoch_train_DAlowLoss[:e + 1], label='Train DAlow loss')
    # l7_4, = plt.plot(t[:e + 1], epoch_test_DAlowLoss[:e + 1], label='Test DAlow loss')
    l7_5, = plt.plot(t[:e + 1], epoch_train_DATLoss[:e + 1], label='Train DATLoss loss')
    # l7_6, = plt.plot(t[:e + 1], epoch_test_DATLoss[:e + 1], label='Test DATLoss loss')
    plt.legend(handles=[l7_1,l7_3,l7_5])
    plt.grid()
    #         plt.gcf().gca().set_ylim(bottom = 0)
    plt.gcf().gca().set_xlim(left=0)
    plt.title('DA Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())


    save = True
    if save:
        plt.figure(num=1)
        plt.savefig('./log/%s/%s/%s-01-loss.png' % (name, time_now, model_type))

        plt.figure(num=2)
        plt.savefig('./log/%s/%s/%s-02-accuracy.png' % (name, time_now, model_type))

        plt.figure(num=3)
        plt.savefig('./log/%s/%s/%s-03-accuracy_per_class.png' % (name, time_now, model_type))

        plt.figure(num=4)
        plt.savefig('./log/%s/%s/%s-04-prec_rec_fmeas.png' % (name, time_now, model_type))

        plt.figure(num=5)
        plt.savefig('./log/%s/%s/%s-05-CELoss.png' % (name, time_now, model_type))

        plt.figure(num=6)
        plt.savefig('./log/%s/%s/%s-06-TLoss.png' % (name, time_now, model_type))

        plt.figure(num=7)
        plt.savefig('./log/%s/%s/%s-07-DAloss.png' % (name, time_now, model_type))
def plotFiguretagert(figure_t_metrics,num_epochs,name,model_type, time_now):
    t = np.linspace(1, num_epochs, num_epochs)
    e=num_epochs
    # print(e,len(figure_train_metrics['rec']),len(figure_test_metrics['ce_loss']))
    epoch_train_nochange_accuracy = figure_t_metrics['nochange_acc']
    epoch_train_change_accuracy = figure_t_metrics['change_acc']
    epoch_train_precision = figure_t_metrics['prec']
    epoch_train_recall = figure_t_metrics['rec']
    epoch_train_Fmeasure = figure_t_metrics['f_meas']
    epoch_train_Loss = figure_t_metrics['Loss']
    epoch_train_ce_loss = figure_t_metrics['ce_loss']
    epoch_train_accuracy = figure_t_metrics['total_acc']
    # epoch_train_focal_loss = figure_train_metrics['focal_loss']
    epoch_train_DALoss = figure_t_metrics['DALoss']
    epoch_train_FeatLoss = figure_t_metrics['FeatLoss']
    # epoch_train_dice_loss = figure_train_metrics['dice_loss']
    epoch_train_Iou = figure_t_metrics['Iou']

    plt.figure(num=1)
    plt.clf()
    train_loss=[]
    test_loss=[]
    # for i in range(len(epoch_train_Loss[:e + 1])):
    #     train_loss.append(epoch_train_clf_loss[:e + 1][i]+epoch_train_marginLoss[:e + 1][i])
    #     test_loss.append(epoch_test_loss[:e + 1][i]+epoch_test_marginLoss[:e + 1][i])

    l1_1, = plt.plot(t[:e + 1], epoch_train_ce_loss[:e + 1], label='Target CE loss')
    l1_3, = plt.plot(t[:e + 1], epoch_train_DALoss[:e + 1], label='Target DA loss')
    l1_5, = plt.plot(t[:e + 1], epoch_train_FeatLoss[:e + 1], label='Target Feat Loss')
    l1_7, = plt.plot(t[:e + 1], epoch_train_Loss, label='Target Total loss')
    plt.legend(handles=[l1_1, l1_3, l1_5, l1_7])
    plt.grid()
    #         plt.gcf().gca().set_ylim(bottom = 0)
    plt.gcf().gca().set_xlim(left=0)
    plt.title('Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=2)
    plt.clf()
    l2_1, = plt.plot(t[:e + 1], epoch_train_accuracy[:e + 1], label='Target accuracy')
    plt.legend(handles=[l2_1])
    plt.grid()
    plt.gcf().gca().set_ylim(0, 1)
    plt.title('Accuracy')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=3)
    plt.clf()
    l3_1, = plt.plot(t[:e + 1], epoch_train_nochange_accuracy[:e + 1], label='Target accuracy: no change')
    l3_2, = plt.plot(t[:e + 1], epoch_train_change_accuracy[:e + 1], label='Target accuracy: change')
    plt.legend(loc='best', handles=[l3_1, l3_2, ])
    plt.grid()
    plt.gcf().gca().set_ylim(0, 1)
    plt.title('Accuracy per class')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=4)
    plt.clf()
    l4_1, = plt.plot(t[:e + 1], epoch_train_precision[:e + 1], label='Target precision')
    l4_2, = plt.plot(t[:e + 1], epoch_train_recall[:e + 1], label='Target recall')
    l4_3, = plt.plot(t[:e + 1], epoch_train_Fmeasure[:e + 1], label='Target Dice/F1')
    l4_7, = plt.plot(t[:e + 1], epoch_train_Iou[:e + 1], label='Target Iou')
    plt.legend(loc='best', handles=[l4_1, l4_2, l4_3,l4_7])
    plt.grid()
    plt.gcf().gca().set_ylim(0, 1)
    #         plt.gcf().gca().set_ylim(bottom = 0)
    #         plt.gcf().gca().set_xlim(left = 0)
    plt.title('Precision, Recall , F-measure and Iou')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=5)
    plt.clf()
    l5_1, = plt.plot(t[:e + 1], epoch_train_ce_loss[:e + 1], label='Target CELoss')
    plt.legend(loc='best', handles=[l5_1])
    plt.grid()
    plt.gcf().gca().set_xlim(left=0)
    plt.title('CE Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())

    plt.figure(num=6)
    plt.clf()
    l6_1, = plt.plot(t[:e + 1], epoch_train_Loss[:e + 1], label='Target Loss')
    plt.legend(loc='best', handles=[l6_1])
    plt.grid()
    plt.gcf().gca().set_xlim(left=0)
    plt.title('Total Loss')
    display.clear_output(wait=True)
    display.display(plt.gcf())


    save = True
    if save:
        plt.figure(num=1)
        plt.savefig('./log/%s/%s/%s-01-lossT.png' % (name, time_now, model_type))

        plt.figure(num=2)
        plt.savefig('./log/%s/%s/%s-02-accuracyT.png' % (name, time_now, model_type))

        plt.figure(num=3)
        plt.savefig('./log/%s/%s/%s-03-accuracy_per_classT.png' % (name, time_now, model_type))

        plt.figure(num=4)
        plt.savefig('./log/%s/%s/%s-04-prec_rec_fmeasT.png' % (name, time_now, model_type))

        plt.figure(num=5)
        plt.savefig('./log/%s/%s/%s-05-CELossT.png' % (name, time_now, model_type))

        plt.figure(num=6)
        plt.savefig('./log/%s/%s/%s-06-TLossT.png' % (name, time_now, model_type))
def MakeRecordFloder(name, time_now,opt,filename_main,pretrain,saveroot):
    os.makedirs('./log/{}/{}'.format(name, time_now))
    os.makedirs('./log/{}/{}/savemodel'.format(name, time_now))
    if pretrain:
        copyfile(saveroot+'/_Training_Log.txt', './log/{}/{}/'.format(name, time_now) + '_Training_Log.txt')
        cfg.TRAINLOG.LOGTXT = open('./log/{}/{}/'.format(name, time_now) + '_Training_Log.txt', 'a')
        print('pretrain log path:', './log/{}/{}'.format(name, time_now))
        print('================ Target Test (%s) ================\n' % time.strftime("%c"))

        cfg.TRAINLOG.LOGTXT.write('\n \n############================ Pretrain (%s) ================############ \n\n' % time.strftime("%c") + '\n')
        cfg.TRAINLOG.EXCEL_LOG = openpyxl.load_workbook(saveroot + "/log.xlsx")
    else:
        cfg.TRAINLOG.LOGTXT = open('./log/{}/{}/'.format(name, time_now) + '_Training_Log.txt', 'w')
        print('log path:', './log/{}/{}'.format(name, time_now))
        cfg.TRAINLOG.EXCEL_LOGDetail= op.Workbook()
        cfg.TRAINLOG.EXCEL_LOGDetail.active
        for i in cfg.TRAINLOG.EXCEL_LOGSheetDetail:
            ws = cfg.TRAINLOG.EXCEL_LOGDetail.create_sheet(i)
            title=['uchgNumF','chgNumF','uchgNumT','chgNumT','uchgDF', 'chgDF', 'DMeanUF', 'DMeanCF','uchgDWF', 'chgDWF',
                   'uchgDT', 'chgDT', 'uchgDWT', 'chgDWT', 'uchgPF', 'chgPF', 'uchgPT','chgPT', 'uchgEF'
                    , 'chgEF', 'EFmean', 'uchgET', 'chgET', 'ETmean','AccOut', 'uchgAccOut', 'chgAccOut', 'mf1Out','AccCI','uchgAccCI', 'chgAccCI', 'mf1CI'
                   ,'AccPre', 'uchgAccPre', 'chgAccPre', 'mf1Pre']
            ws.append(
                title
                )
            orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
            green_fill = PatternFill(fill_type='solid', fgColor="51DD54")
            blue_fill = PatternFill(fill_type='solid', fgColor="3EF0F0")
            red_fill = PatternFill(fill_type='solid', fgColor="FB3341")
            yellow_fill = PatternFill(fill_type='solid', fgColor="DFF226")
            j=1
            for T in title:
                if 'F' in T:
                    ws.cell(row=1, column=j).fill = orange_fill
                elif 'T' in T:
                    ws.cell(row=1, column=j).fill = green_fill
                elif 'C' in T:
                    ws.cell(row=1, column=j).fill = blue_fill
                elif 'O' in T:
                    ws.cell(row=1, column=j).fill = red_fill
                elif 'P' in T:
                    ws.cell(row=1, column=j).fill = yellow_fill
                j=j+1
        # excel
        cfg.TRAINLOG.EXCEL_LOG = op.Workbook()
        cfg.TRAINLOG.EXCEL_LOG.active
        for i in cfg.TRAINLOG.EXCEL_LOGSheet:

            ws = cfg.TRAINLOG.EXCEL_LOG.create_sheet(i)
            ws.append(
                ['type', 'n_epochs', 'ce_loss', 'DAlow','DAEntropy','DATLoss', 'FeatLoss', 'Acc', 'nochange_acc', 'change_acc', 'recall'
                , 'Fmeasure','fm1', 'IOU', 'prec', 'TN', 'TP', 'FN', 'FP','accT','unchgaccT','chgaccT','mf1T'])
        worksheets = cfg.TRAINLOG.EXCEL_LOG.sheetnames
        # cfg.TRAINLOG.EXCEL_LOG.get_sheet_by_name()
        # print('worksheetsworksheets',worksheets)
    cfg.TRAINLOG.LOGTXT.write('log path:'+ './log/{}/{}'.format(name, time_now) + '\n')
    # cfg.TRAINLOG.LOGJSON = open('./log/{}/{}/'.format(name, time_now) + '_Training_Log.json', 'w')
    # json.dump('Training_DATASET: ' +cfg.TRAINLOG.DATA_NAMES[opt.s] +
    #           '\\n', cfg.TRAINLOG.LOGJSON)
    cfg.TRAINLOG.LOGTXT.write('Training_DATASET: ' +
                              cfg.TRAINLOG.DATA_NAMES[opt.s] + '\n')
    targetName=[cfg.TRAINLOG.DATA_NAMES[i] for i in range(len(cfg.TRAINLOG.DATA_NAMES)) if i!=opt.s]
    cfg.TRAINLOG.LOGTXT.write('Target_DATASET: ' +str(targetName) + '\n')
    # json.dump('Target_DATASET: ' +str(targetName) + '\n', cfg.TRAINLOG.LOGJSON)

    shutil.copytree('./model/', './log/{}/{}/models'.format(name, time_now))
    shutil.copytree('./util/', './log/{}/{}/utils'.format(name, time_now))
    shutil.copytree('./data/', './log/{}/{}/data'.format(name, time_now))
    shutil.copytree('./DCNv2/', './log/{}/{}/DCNv2'.format(name, time_now))
    shutil.copytree('./option/', './log/{}/{}/option'.format(name, time_now))
    shutil.copytree('./modelDA/', './log/{}/{}/modelDA'.format(name, time_now))
    shutil.copytree('./predictions/', './log/{}/{}/predictions'.format(name, time_now))

    newtargetpath = './log/{}/{}/'.format(name, time_now) + filename_main
    shutil.copyfile(filename_main, newtargetpath)


    #checkpoint
    # cfg.TRAINLOG.ITER_PATH = os.path.join('./log/{}/{}/savemodel'.format(name, time_now), 'iter.txt')
    # if opt.load_pretrain:
    #     try:
    #         start_epoch, epoch_iter = np.loadtxt(cfg.TRAINLOG.ITER_PATH, delimiter=',', dtype=int)
    #     except:
    #         start_epoch, epoch_iter = 1, 0
    #     print('Resuming from epoch %d at iteration %d' % (start_epoch, epoch_iter))
    # else:
    #     start_epoch, epoch_iter = 1, 0
    start_epoch, epoch_iter = 1, 0
    return start_epoch, epoch_iter

def confuseMatrix(pred,label_source):
    pred = pred.cpu().numpy()
    label_source = label_source.cpu().numpy()
    pr = (pred > 0)
    gt = (label_source > 0)
    # print(pr)
    tp_e = np.logical_and(pr, gt).sum()
    tn_e = np.logical_and(np.logical_not(pr), np.logical_not(gt)).sum()
    fp_e = np.logical_and(pr, np.logical_not(gt)).sum()
    fn_e = np.logical_and(np.logical_not(pr), gt).sum()

    return tp_e,tn_e,fp_e,fn_e
def init_method(net, init_type='normal'):
    def init_func(m):
        classname = m.__class__.__name__
        if hasattr(m, 'resnet'):
            pass
        elif hasattr(m, 'weight') and (classname.find('Conv') != -1 or classname.find('Linear') != -1):
            if init_type == 'normal':
                init.normal_(m.weight.data, 0.0, 0.02)
            elif init_type == 'xavier':
                init.xavier_normal_(m.weight.data, gain=0.02)
            elif init_type == 'kaiming':
                init.kaiming_normal_(m.weight.data, a=0, mode='fan_in')
            elif init_type == 'orthogonal':
                init.orthogonal_(m.weight.data, gain=0.02)
            else:
                raise NotImplementedError('initialization method [%s] is not implemented' % init_type)
            if hasattr(m, 'bias') and m.bias is not None:
                init.constant_(m.bias.data, 0.0)
        elif classname.find('BatchNorm2d') != -1:
            init.normal_(m.weight.data, 1.0, 0.02)
            init.constant_(m.bias.data, 0.0)

    print('initialize network with %s' % init_type)
    net.apply(init_func)